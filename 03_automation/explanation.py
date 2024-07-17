import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the workbook
wb = xl.load_workbook('transactions.xlsx')

# Select the sheet
sheet = wb['Sheet1']

# Access a specific cell
cell = sheet['A1']  # It's better to use uppercase letters

# Alternatively, access the cell using row and column indices
cell = sheet.cell(1, 1)

# Print the value of the cell
# print(cell.value)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    modified_price = cell.value * 0.9
    modified_price_cell = sheet.cell(row, 4)
    modified_price_cell.value = modified_price
  
    
values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e2")

wb.save('refined_transactions.xlsx')