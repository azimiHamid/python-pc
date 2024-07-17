import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(input_filename, output_filename):
    wb = xl.load_workbook(input_filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        modified_price = cell.value * 0.9
        modified_price_cell = sheet.cell(row, 4)
        modified_price_cell.value = modified_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(output_filename)

input_filename = "transactions.xlsx"
output_filename = "refined_transactions.xlsx"
process_workbook(input_filename, output_filename)